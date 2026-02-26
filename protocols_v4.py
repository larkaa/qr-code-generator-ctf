#!/usr/bin/env python3
"""
Import all CSV files from a given directory into a single Excel workbook.
Each CSV becomes a worksheet named after the file (without extension) and
receives the formatting described in the prompt.

Usage:
    python main.py /path/to/csv_folder output.xlsx
"""

import argparse
import pathlib
import sys
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------
# Styles
# ----------------------------------------------------------------------
BLACK_BG   = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
PURPLE_BG  = PatternFill(start_color="800080", end_color="800080", fill_type="solid")
GRAY_BG    = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
YELLOW_BG  = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF")
RED_FONT   = Font(color="FF0000")
GREEN_FONT = Font(color="00B050")


# Define search patterns as (category_name, [regex_patterns])
# All port-number patterns use lookarounds so they match at string boundaries too.
SEARCH_PATTERNS = [
    # forbidden protocols
    ("(f) Rexec, exec, kexe port 512", [r'r?k?exec', r'[^0-9]512[^0-9]', r'tcp-?_?512']), 
    ("(f) FTP port 20/21", [r'\bftp\b', r'(?:tcp|udp)[_-]?2[01]\b', r'(?:port|:)2[01]\b']),
    ("(f) TFTP port 69", [r'\btfpt\b', r'[^0-9.]69[^0-9.]', r'udp[_-]?69\b'] ),
    ("(f) Login/rlogin/klogin port 513", [r'r?k?login\b', r'(?:tcp|udp)[_-]?513\b', r'(?:tcp|udp)[_-]?543\b']),
    ("(f) Netbios ports 137-139", [r'netbios', r'[^0-9.]13[789][^0-9.]', r'tcp-?_?13[789]\b']),
    ("(f) NFS v1-3, port 111, 2049", [ r'\bnfs', r'[^0-9]2049\b', r'[^0-9]111\b', ]), 
    ("(f) RDP standard port 3389", [r'(?:rdp|remote[_\-\s]desktop)', r'(?:tcp|udp)[_-]?3389\b']),
    ("(f) SMBv1 cifs port 445", [  r'\bsmb\b', r'smb_v1', r'\bcifs', r'[^0-9]445\b',  r'(?:tcp|udp)[_-]?445\b']) 
    ("(f) SQLnet v1 port 1525", [ r'sqlnet', r'[^0-9]1525[^0-9]'   ]),
    ("(f) SSH v1", [r'[^a-z]ssh(?![ _]version[ _]2)', r'(?:tcp|udp)[_-]?22\b', r'(?:port|:)22\b', r'ssh_version_1\b']),   
    ("(f) Telnet port 23", [r'telnet',   r'(?:tcp|udp)[_-]?23\b', r'(?:port|:)23\b'  ]),  
    ("(f) VNC / remote services ports 5500,5800", [r'remote', r'vnc', r'[^0-9]5500[^0-9]', r'[^0-9]5800\b']), 
    ("(f) x11 ports 6000-6063", [r'x11',  r'[^0-9.]60[0-6][0-9][^0-9.]'  ]), # 6000 - 6063
    ("(f) snmp v1,2 ports 161/2", [ r'snmp', r'[^0-9]16[12]\b', ]),  
     
    # restricted protocols
    ("(r) HTTP port 80", [r'http[^s\w]', r'(?:tcp|udp)[_-]?80\b', r'(?:port|:)80\b']),
    ("(r) MySQL / MariaDB ports 3306/7", [r'\bmysql\b', r'\bmaria(?:db)?\b', r'(?:tcp|udp)[_-]?330[67]\b']),
    ("(r) Oracle-Listener", [r'oracle', r'[^0-9]1521[^0-9]', r'tcp[_-]?1521[^0-9]']), 
    ("(r) SQLServer ports 1433/4", [r'(?:ms[_-]?sql|mssql|sql[_-]?server)\b', r'(?:tcp|udp)[_-]?143[34]\b']),
    ("(r) SQLServer broker port [457]022",    [ r'[457]022\b']), # sql broker, mirror 
    ("(r) SQLServer browser port 2382",    [ r'2382\b'   ]), # sql server browser
    ("(r) PostgreSQL port 5432", [r'postgre[^s]', r'[^0-9]5432\b', r'tcp[_-]?5432\b']), 
    ("(r) RPC port 135", [r'\brpc\b', r'(?:tcp|udp)[_-]?135\b']),
    ("(r) LDAP port 389", [r'\bldap\b', r'[^0-9]389\b', r'tcp[_-]?389\b']), 
]


# Columns to search for protocol matches
COLUMNS_TO_SEARCH = ["Name", "Source", "Destination", "Services & Applications"]


def categorize(df, search_patterns=SEARCH_PATTERNS, columns_to_search=COLUMNS_TO_SEARCH):
    """
    Add '--Forbidden--' and '--Restricted--' columns to the DataFrame.
    Matches are only checked against the specified target columns.
    """
    for_column = "--Forbidden--"
    res_column = "--Restricted--"

    df[for_column] = ""
    df[res_column] = ""

    # Compile all regex patterns up front
    compiled_patterns = [
        (category, [re.compile(p, re.IGNORECASE) for p in patterns])
        for category, patterns in search_patterns
    ]

    # Determine which target columns actually exist in this DataFrame
    available_cols = [c for c in columns_to_search if c in df.columns]

    for idx, row in df.iterrows():
        # Build search string from target columns only
        row_str = " ".join(str(row[col]) for col in available_cols if pd.notna(row[col]))

        # Skip section header rows
        if row_str.lower().startswith('section'):
            continue

        matched_for = set()
        matched_res = set()

        for category, patterns in compiled_patterns:
            for pattern in patterns:
                if pattern.search(row_str):
                    if category.startswith("(f)"):
                        matched_for.add(category)
                    elif category.startswith("(r)"):
                        matched_res.add(category)
                    break  # one match per category is enough

        if matched_for:
            df.at[idx, for_column] = ", ".join(sorted(matched_for))
        elif matched_res:
            df.at[idx, res_column] = ", ".join(sorted(matched_res))

    return df


def format_sheet(ws, header_row=1):
    """Apply all conditional formatting to a worksheet."""
    # Header row: black background, white text
    for cell in ws[header_row]:
        cell.fill = BLACK_BG
        cell.font = WHITE_FONT

    # Map lowercase header names to 1-based column indices
    col_idx = {
        cell.value.strip().lower(): cell.column
        for cell in ws[header_row]
        if cell.value
    }

    src_col  = col_idx.get("source")
    dst_col  = col_idx.get("destination")
    svc_col  = col_idx.get("services & applications")
    act_col  = col_idx.get("action")
    rule_col = col_idx.get("type")

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        # Section rows: purple background
        if rule_col:
            rule_val = row[rule_col - 1].value
            if isinstance(rule_val, str) and rule_val.strip().lower() == "section":
                for cell in row:
                    cell.fill = PURPLE_BG
                    cell.font = WHITE_FONT
                continue

        # Disabled rows: gray background
        if rule_col:
            rule_val = row[rule_col - 1].value
            if isinstance(rule_val, str) and "disabled" in rule_val.lower():
                for cell in row:
                    cell.fill = GRAY_BG

        # "Any" in Source / Destination / Services: yellow background
        for col in (src_col, dst_col, svc_col):
            if col:
                val = row[col - 1].value
                # Guard against None/non-string values
                if isinstance(val, str) and "any" in val.lower():
                    row[col - 1].fill = YELLOW_BG

        # Action column: red for Drop, green for Accept
        if act_col:
            val = row[act_col - 1].value
            if val == "Drop":
                row[act_col - 1].font = RED_FONT
            elif val == "Accept":
                row[act_col - 1].font = GREEN_FONT


def main():
    parser = argparse.ArgumentParser(
        description="Combine CSVs into one Excel workbook with formatting."
    )
    parser.add_argument("csv_dir", type=pathlib.Path,
                        help="Directory containing CSV files")
    parser.add_argument("output", type=pathlib.Path,
                        help="Path for the generated .xlsx file")
    args = parser.parse_args()

    if not args.csv_dir.is_dir():
        sys.exit(f"Error: {args.csv_dir} is not a directory")

    wb = Workbook()
    wb.remove(wb.active)  # Remove default empty sheet

    csv_files = sorted(args.csv_dir.glob("*.csv"))
    if not csv_files:
        sys.exit("No CSV files found in the supplied directory.")

    for csv_path in csv_files:
        # FIX: Do NOT call reset_index() — it inserts a spurious 'index' column
        # that shifts all original columns right by one, breaking column lookups.
        # index_col=False forces pandas to treat ALL columns as data columns.
        # index_col=None (the default) allows pandas to silently promote the
        # first numeric-looking column (e.g. "No.") to the row index, which
        # drops it from df.columns and shifts everything else left by one.
        df = pd.read_csv(csv_path, index_col=False)
        df = categorize(df)

        sheet_name = csv_path.stem[:31]  # Excel sheet name limit = 31 chars
        ws = wb.create_sheet(title=sheet_name)

        # Write header row (row 1)
        for c_idx, col_name in enumerate(df.columns, start=1):
            ws.cell(row=1, column=c_idx, value=col_name)

        # Write data rows
        for r_idx, row in enumerate(df.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Auto-filter on header row
        ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}1"

        # Freeze first row and first column
        ws.freeze_panes = "B2"

        # Apply formatting
        format_sheet(ws)

    wb.save(args.output)
    print(f"Workbook saved to {args.output}")


if __name__ == "__main__":
    main()
